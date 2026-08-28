from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

import httpx
from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader

from ..core.config import Settings
from ..domain.models import AttachmentExtraction
from ..planning.llm import deepseek_complete


async def extract_attachment(
    file_id: str,
    path: Path,
    mime_type: str,
    settings: Settings,
) -> AttachmentExtraction:
    warnings: list[str] = []
    text = ""
    if mime_type == "application/pdf":
        text = "\n".join(page.extract_text() or "" for page in PdfReader(path).pages)
    elif mime_type.endswith("wordprocessingml.document"):
        text = "\n".join(paragraph.text for paragraph in Document(path).paragraphs)
    elif mime_type.endswith("spreadsheetml.sheet"):
        workbook = load_workbook(path, read_only=True, data_only=True)
        text = "\n".join(
            " | ".join(str(value) for value in row if value is not None)
            for sheet in workbook.worksheets
            for row in sheet.iter_rows(values_only=True)
        )
    elif mime_type in {"text/markdown", "text/plain"}:
        text = path.read_text(encoding="utf-8", errors="replace")

    # Attachments are semantic input. Do not infer places or dates from local
    # keyword lists: without the Agent we preserve only a safe text preview.
    extracted = AttachmentExtraction(
        file_id=file_id,
        text_preview=text.strip()[:1000],
    )
    if settings.deepseek_api_key or settings.ollama_api_key:
        llm = await _extract_with_deepseek(path, mime_type, text, settings)
        if llm:
            for field in ("places", "hotels", "dates", "order_numbers"):
                incoming = [str(item) for item in llm.get(field, [])]
                if field in {"places", "hotels"}:
                    incoming = [_clean_place_name(item) for item in incoming]
                setattr(extracted, field, list(dict.fromkeys(incoming))[:50])
            if not extracted.text_preview or llm.get("summary"):
                extracted.text_preview = str(llm.get("summary") or "")[:1000]
        elif mime_type.startswith("image/"):
            warnings.append("图像识别服务未返回结构化结果，请人工补充或重试")
    else:
        warnings.append("未配置附件理解 Agent，仅保留文本预览，未进行语义猜测")
    extracted.warnings.extend(warnings)
    return extracted


def _clean_place_name(value: str) -> str:
    return " ".join(value.split())


async def _extract_with_deepseek(
    path: Path,
    mime_type: str,
    text: str,
    settings: Settings,
) -> dict[str, Any] | None:
    prompt = (
        "你是 RoadMan 附件信息提取器。只提取，不规划，不猜测。"
        "返回单个 JSON：places,hotels,dates,order_numbers 均为字符串数组，summary 为短摘要。"
        "无法确认的字段返回空数组。"
    )
    attachment_text = text[:12000]
    if mime_type.startswith("image/"):
        # deepseek-v4-flash is a text model; do not send unsupported image
        # bytes.  Keep the extraction explicit so the caller can supplement
        # the missing visual fields manually.
        attachment_text = "[图片附件：当前文本模型无法可靠读取图片，请人工补充地点、日期或订单信息。]"
    request_prompt = f"{prompt}\n附件文本：{attachment_text}"
    try:
        content = await deepseek_complete(
            settings,
            request_prompt,
            timeout=settings.deepseek_timeout_seconds,
        )
        match = re.search(r"\{.*\}", content, re.DOTALL)
        value = json.loads(match.group(0)) if match else None
        return value if isinstance(value, dict) else None
    except (httpx.HTTPError, ValueError, TypeError, json.JSONDecodeError):
        return None


# Compatibility name used by existing API tests and local integrations.  It
# now delegates to DeepSeek and never calls the former Ollama endpoint.
_extract_with_ollama = _extract_with_deepseek
